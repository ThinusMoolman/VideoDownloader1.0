from tkinter import *
from tkinter import ttk
from time import sleep
import openpyxl as xl

root = Tk()
frame = Frame(root)
frame.pack(side="top", expand=True, fill="both")
Folder_Path = StringVar()
wb = ""
sheet = ""
column_choice = ""
choice = ""
choices = ""
Folder = ""


def QuitWindow():
    sys.exit()


def Home_Page():
    for widget in frame.winfo_children():
        widget.destroy()
    frame.pack_forget()
    welcome_text = '''Welcome user!
    This is free software made by Thinus Moolman
    to make your life easier to sort all your excel files!
    If you are ready, click the button below to begin!'''
    DescriptionLabel = Label(frame, text=welcome_text, font=("Agency FB", 45), fg="grey")
    DescriptionLabel.grid()
    frame_but = Button(frame, text="Start", font=("Agency FB", 45), bg="grey", fg="white", command=lambda: Page1())
    frame_but.grid()
    frame.pack()


def Page1():
    global Folder_Path
    for widget in frame.winfo_children():
        widget.destroy()
    frame.pack_forget()
    text = '''Okay user so first we'll have to
    open your excel file so
    please paste the path of it!'''
    ExplainLabel2 = Label(frame, text=text, font=("Agency FB", 45), fg="grey")
    ExplainLabel2.grid()
    Entry(frame, width=50, textvariable=Folder_Path).grid()
    Label(frame, text="Alright! Now press this button to continue!", fg="grey", font=("Agency FB", 45)).grid()
    Button(frame, text="Continue", command=lambda: Page2(), fg="grey").grid()
    frame.pack()


def Page2():
    global Folder_Path
    global wb
    global choice
    global choices
    global Folder
    font = ("Agency FB", 45)
    for widget in frame.winfo_children():
        widget.destroy()
    frame.pack_forget()
    Label(frame, text="Let's try to open your file now!", font=font, fg="grey").grid()
    sleep(4)
    Folder = Folder_Path.get()
    print(Folder)
    try:
        wb = xl.load_workbook(Folder)
        Label(frame, text="Success", font=font, fg="green").grid()
        Label(frame, text="Great so let's get started with the sorting!", font=font, fg="grey").grid()
        Label(frame, text="The different options are displayed below: ").grid()
        choices = ["Alphabetical sort", "Numerical sort"]
        choose_choices = ttk.Combobox(frame, values=choices, font=("Agency FB", 25))
        choose_choices.grid()

        def ColumnChooser():
            global column_choice
            global sheet
            global choice
            global choices
            choice = choose_choices.get()
            print(choice)
            if choice == choices[0]:
                sheet = wb["Sheet1"]
                column_choices = []
                maxColumn = sheet.max_column
                x = 0
                for i in range(1, maxColumn):
                    column_choices.append(i)
                Label(frame, text="Choose which row to sort", font=font, fg="Green").grid()
                column_choice = ttk.Combobox(frame, values=column_choices, font=("Agency FB", 25))
                column_choice.grid()
                Button(frame, text="Continue", font=font, fg="Green", command=lambda: Page3()).grid()
            if choice == choices[1]:
                sheet = wb["Sheet1"]
                column_choices = []
                maxColumn = sheet.max_column
                x = 0
                for i in range(1, maxColumn):
                    column_choices.append(i)
                Label(frame, text="Choose which row to sort", font=font, fg="Green").grid()
                column_choice = ttk.Combobox(frame, values=column_choices, font=("Agency FB", 25))
                column_choice.grid()
                Button(frame, text="Continue", font=font, fg="Green", command=lambda: Page3()).grid()
                print("test")

        Button(frame, text="Continue", font=font, fg="Green", command=lambda: ColumnChooser()).grid()
    except Exception as e:
        Label(frame, text=f"Error: {e}", fg="red").grid()
        Label(frame, text="Would you like to restart?", fg="red", font=font).grid()
        Button(frame, text="Restart", fg="Green", font=font, command=lambda: Home_Page()).grid()
        Button(frame, text="Quit", fg="Red", font=font, command=lambda: QuitWindow()).grid()
    frame.pack()


def Page3():
    global wb
    global choice
    global choices
    global column_choice
    global sheet
    global Folder
    Column_choose = column_choice.get()
    for widget in frame.winfo_children():
        widget.destroy()
    frame.pack_forget()
    name_dictionary = {}
    if choice == choices[0]:
        for i in range(1, sheet.max_row):
            if Column_choose == "1":
                Value1 = sheet.cell(i, 1).value
                Value2 = sheet.cell(i, 2).value
                name_dictionary[Value1] = Value2
            if Column_choose == "2":
                Value1 = sheet.cell(i, 1).value
                Value2 = sheet.cell(i, 2).value
                name_dictionary[Value2] = Value1

        print(name_dictionary)
        name_dictionary_sorted = sorted(name_dictionary)
        x = 1
        print(name_dictionary_sorted)
        for y in name_dictionary_sorted:
            sheet.cell(x, 1).value = y
            sheet.cell(x, 2).value = name_dictionary[y]
            x += 1
        wb.save("D:\Sorted.xlsx")
        Label(text="SUCCESS", font=("Agency FB", 70), fg="green").grid()
        Label(text="Your file has been saved to D:\Sorted.xlsx", font=("Agency FB", 45), fg="green").grid()
        Button(text="Sort another", font=("Agency FB", 25), fg="green", command=lambda: Page1()).grid()
        Button(text="Quit", font=("Agency FB", 25), fg="red", command=lambda: QuitWindow()).grid()
    if choice == choices[1]:
        try:
            for i in range(1, sheet.max_row):
                if Column_choose == "1":
                    Value1 = sheet.cell(i, 1).value
                    Value2 = sheet.cell(i, 2).value
                    name_dictionary[Value2] = int(Value1)
                if Column_choose == "2":
                    Value1 = sheet.cell(i, 1).value
                    Value2 = sheet.cell(i, 2).value
                    name_dictionary[Value1] = int(Value2)

            print(name_dictionary)
            name_dictionary_sorted = sorted(name_dictionary, key=name_dictionary.get)
            x = 1
            print(name_dictionary_sorted)
            for y in name_dictionary_sorted:
                sheet.cell(x, 1).value = y
                sheet.cell(x, 2).value = name_dictionary[y]
                x += 1
            wb.save("D:\Sorted.xlsx")
            Label(text="SUCCESS", font=("Agency FB", 70), fg="green").grid()
            Label(text="Your file has been saved to D:\Sorted.xlsx", font=("Agency FB", 45), fg="green").grid()
            Button(text="Sort another", font=("Agency FB", 25), fg="green", command=lambda: Page1()).grid()
            Button(text="Quit", font=("Agency FB", 25), fg="red", command=lambda: QuitWindow()).grid()
        except Exception as e:
            Label(text=f"Error {e}", fg="red").grid()
            Button(text="Try again", font=("Agency FB", 25), fg="green", command=lambda: Page1()).grid()
            Button(text="Quit", font=("Agency FB", 25), fg="red", command=lambda: QuitWindow()).grid()


Home_Page()
root.mainloop()
